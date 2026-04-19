"""
Generate a Jira-style Excel project plan for the 4-week AI Sales Pipeline build.
Multiple sheets: Project Overview, Sprint Backlog, Daily Schedule, Kanban Board,
Milestones, Dependencies, and Risk Register.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date, timedelta

wb = Workbook()

# ─── Color palette ───
INDIGO = "4F46E5"
DARK_INDIGO = "312E81"
PURPLE = "7C3AED"
TEAL = "0D9488"
GREEN = "16A34A"
YELLOW = "CA8A04"
RED = "DC2626"
ORANGE = "EA580C"
GRAY = "64748B"
LIGHT_GRAY = "F1F5F9"
WHITE = "FFFFFF"
LIGHT_INDIGO = "EEF2FF"
LIGHT_PURPLE = "F5F3FF"
LIGHT_TEAL = "F0FDFA"
LIGHT_GREEN = "F0FDF4"
LIGHT_YELLOW = "FEFCE8"
LIGHT_RED = "FEF2F2"
LIGHT_ORANGE = "FFF7ED"

# ─── Helpers ───
thin_border = Border(
    left=Side(style='thin', color="E2E8F0"),
    right=Side(style='thin', color="E2E8F0"),
    top=Side(style='thin', color="E2E8F0"),
    bottom=Side(style='thin', color="E2E8F0"),
)

def style_header(cell, bg=INDIGO, fg=WHITE, size=11):
    cell.font = Font(bold=True, color=fg, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

def style_cell(cell, bold=False, color="1E1E2E", size=10, wrap=True, align="left", bg=None):
    cell.font = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin_border
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title_row(ws, title, subtitle, row=1, cols=10, color=DARK_INDIGO):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, color=color, size=20, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 32

    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=cols)
    c2 = ws.cell(row=row+1, column=1, value=subtitle)
    c2.font = Font(color=GRAY, size=11, italic=True, name="Calibri")
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row+1].height = 22

# ═══════════════════════════════════════════════════════════════════
#  SHEET 1: PROJECT OVERVIEW
# ═══════════════════════════════════════════════════════════════════

ws = wb.active
ws.title = "Project Overview"

add_title_row(ws, "AnalyticsGear AI Sales Pipeline", "4-Week Sprint Plan · Jira-Style Project Tracker", cols=6)

# Project info
ws.cell(row=4, column=1, value="Project Information").font = Font(bold=True, size=13, color=DARK_INDIGO)
info = [
    ("Project Code", "AG-AISP"),
    ("Project Name", "AI-Native Sales Pipeline"),
    ("Project Lead", "Imroz (Founder)"),
    ("Start Date", "Monday, April 20, 2026"),
    ("Target Go-Live", "Sunday, May 17, 2026"),
    ("Total Duration", "4 weeks / 28 days"),
    ("Sprints", "4 one-week sprints"),
    ("Total Story Points", "108 tasks across 6 epics"),
    ("Daily Time Budget", "4-6 hours/day"),
    ("Methodology", "Agile / Scrum-inspired, solo dev"),
]

for i, (k, v) in enumerate(info, start=5):
    c1 = ws.cell(row=i, column=1, value=k)
    c2 = ws.cell(row=i, column=2, value=v)
    style_cell(c1, bold=True, color=GRAY, align="right")
    style_cell(c2, bold=True, color=DARK_INDIGO, align="left")

# Epics section
ws.cell(row=16, column=1, value="Epics").font = Font(bold=True, size=13, color=DARK_INDIGO)

epic_headers = ["Epic Code", "Epic Name", "Description", "Sprint", "Tasks", "Priority"]
for i, h in enumerate(epic_headers, 1):
    style_header(ws.cell(row=17, column=i, value=h))

epics = [
    ("EP-01", "Foundation & Setup", "Project setup, environments, Qdrant, Claude API integration", "Sprint 1", 12, "P0"),
    ("EP-02", "RAG + Vector Knowledge Base", "Embedding pipeline, vector collections, RAG retrieval, knowledge ingestion", "Sprint 1", 15, "P0"),
    ("EP-03", "MCP Server Layer", "6 MCP servers: CRM, VectorDB, Email, Scraper, Apollo, LinkedIn", "Sprint 1-2", 24, "P0"),
    ("EP-04", "AI Agents", "Prospecting, Outreach, Follow-up, Analytics agents with ReAct pattern", "Sprint 2-3", 28, "P0"),
    ("EP-05", "Memory & Feedback Loop", "Lead memory, pattern memory, self-improvement mechanisms", "Sprint 3", 12, "P1"),
    ("EP-06", "Production & Go-Live", "Deployment, monitoring, email warmup completion, launch", "Sprint 4", 17, "P0"),
]

for i, ep in enumerate(epics, start=18):
    for j, v in enumerate(ep, 1):
        cell = ws.cell(row=i, column=j, value=v)
        style_cell(cell, wrap=True, align="left")
        if j == 1:
            style_cell(cell, bold=True, color=PURPLE, align="center")
        if j == 6:  # Priority
            if v == "P0":
                style_cell(cell, bold=True, color=RED, align="center", bg=LIGHT_RED)
            elif v == "P1":
                style_cell(cell, bold=True, color=ORANGE, align="center", bg=LIGHT_ORANGE)
    ws.row_dimensions[i].height = 32

# Sprint summary
ws.cell(row=25, column=1, value="Sprint Summary").font = Font(bold=True, size=13, color=DARK_INDIGO)

sprint_headers = ["Sprint", "Dates", "Goal", "Key Deliverable", "Tasks"]
for i, h in enumerate(sprint_headers, 1):
    style_header(ws.cell(row=26, column=i, value=h), bg=PURPLE)

sprints = [
    ("Sprint 1", "Apr 20 – Apr 26", "Foundation + Knowledge Base + First MCP Server",
     "Working RAG pipeline retrieving company data + blog content. mcp-crm server operational.", 27),
    ("Sprint 2", "Apr 27 – May 3", "All MCP Servers + Prospecting Agent Live",
     "All 6 MCP servers running. Prospecting Agent autonomously finding 25 qualified leads/day.", 27),
    ("Sprint 3", "May 4 – May 10", "Outreach + Follow-up + Memory",
     "Full multi-channel outreach with RAG personalization. Memory-augmented follow-ups.", 26),
    ("Sprint 4", "May 11 – May 17", "Analytics + Production + Go-Live",
     "System live in production. Daily pipeline running. First real outreach sent.", 28),
]

for i, s in enumerate(sprints, start=27):
    for j, v in enumerate(s, 1):
        cell = ws.cell(row=i, column=j, value=v)
        style_cell(cell, wrap=True, align="left")
        if j == 1:
            style_cell(cell, bold=True, color=PURPLE, align="center", bg=LIGHT_PURPLE)
    ws.row_dimensions[i].height = 48

# Key Success Metrics
ws.cell(row=33, column=1, value="Definition of Done / Success Criteria").font = Font(bold=True, size=13, color=DARK_INDIGO)
dod = [
    "All 6 MCP servers running and discoverable",
    "All 4 agents operational and executing daily",
    "Vector DB populated with 200+ company profiles + AnalyticsGear content",
    "RAG retrieval returning relevant context with > 85% relevance",
    "Email domain warmed up and deliverability validated (no blacklists)",
    "First 10 real cold emails sent, tracking active",
    "Daily pipeline running autonomously on schedule via cron/Prefect",
    "Monitoring dashboard showing LLM costs, agent success rate, engagement metrics",
    "Runbook documented for common issues",
    "At least 1 production bug survived + fixed in first 48 hours (proves resilience)",
]

for i, d in enumerate(dod, start=34):
    c = ws.cell(row=i, column=1, value=f"✓  {d}")
    style_cell(c, color=GREEN, bold=True)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

set_col_widths(ws, [18, 30, 40, 18, 10, 12])

# ═══════════════════════════════════════════════════════════════════
#  SHEET 2: SPRINT BACKLOG (Jira-style task list)
# ═══════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Sprint Backlog")
add_title_row(ws, "Sprint Backlog", "All 108 tasks across 4 sprints · Jira-style issue tracker", cols=12)

# Headers
backlog_headers = [
    "Task ID", "Sprint", "Day", "Date", "Epic", "Title", "Description",
    "Type", "Priority", "Estimate (hrs)", "Status", "Acceptance Criteria"
]
for i, h in enumerate(backlog_headers, 1):
    c = ws.cell(row=4, column=i, value=h)
    style_header(c, bg=DARK_INDIGO)

# ─── ALL TASKS ───
# Format: (task_id, sprint, day, date, epic, title, description, type, priority, est, status, ac)
TASKS = [
    # ========== SPRINT 1 (Week 1) ==========
    # Day 1 — Mon Apr 20 — Environment Setup
    ("AG-001", 1, 1, "2026-04-20", "EP-01", "Initialize git repo & project structure",
     "Create GitHub repo 'sales-pipeline-ai'. Set up directory structure per blueprint (agents/, mcp_servers/, rag/, memory/, feedback/, config/, prompts/).",
     "Task", "P0", 1, "To Do",
     "Repo initialized with README, .gitignore, LICENSE. Full folder structure created."),
    ("AG-002", 1, 1, "2026-04-20", "EP-01", "Set up Python 3.12 venv",
     "Install Python 3.12, create virtual environment, configure pyproject.toml with dependency groups (core, mcp, agents, rag).",
     "Task", "P0", 0.5, "To Do",
     "Venv active, pyproject.toml validates, pip install runs clean."),
    ("AG-003", 1, 1, "2026-04-20", "EP-01", "Install core dependencies",
     "Install: anthropic, qdrant-client, mcp-sdk, voyageai, gspread, sendgrid, python-dotenv, apscheduler, pydantic, httpx.",
     "Task", "P0", 0.5, "To Do",
     "All packages installed, no conflicts, import smoke test passes."),
    ("AG-004", 1, 1, "2026-04-20", "EP-01", "Configure .env and secrets management",
     "Create .env.example and .env. Add API keys: ANTHROPIC_API_KEY, VOYAGE_API_KEY, APOLLO_API_KEY, SENDGRID_API_KEY, GOOGLE_CREDS_PATH.",
     "Task", "P0", 0.5, "To Do",
     "All keys loaded via python-dotenv. Secrets never committed."),
    ("AG-005", 1, 1, "2026-04-20", "EP-01", "Set up Docker Compose with Qdrant",
     "Create docker-compose.yml with Qdrant service. Mount volume for persistence. Test container starts.",
     "Task", "P0", 1, "To Do",
     "docker compose up starts Qdrant on :6333. UI accessible at localhost:6333/dashboard."),
    ("AG-006", 1, 1, "2026-04-20", "EP-01", "Validate Claude API access",
     "Test Claude API with claude-sonnet-4-6. Verify tool use works. Measure baseline latency and cost.",
     "Task", "P0", 0.5, "To Do",
     "Successful API call, tool use working, costs logged to spreadsheet."),

    # Day 2 — Tue Apr 21 — Vector DB Foundation
    ("AG-007", 1, 2, "2026-04-21", "EP-02", "Design vector collection schemas",
     "Define 6 collections: company_profiles, prospect_contacts, ag_content, email_history, industry_knowledge, successful_patterns. Decide vector size (1024 for voyage-3).",
     "Design", "P0", 1, "To Do",
     "Schema document created. Collection configs checked into config/collections.yaml."),
    ("AG-008", 1, 2, "2026-04-21", "EP-02", "Create all Qdrant collections",
     "Write init_collections.py. Create all 6 collections with proper vector params (COSINE distance, 1024 dims) and payload indexes.",
     "Task", "P0", 1.5, "To Do",
     "All 6 collections visible in Qdrant dashboard. Script idempotent (safe to re-run)."),
    ("AG-009", 1, 2, "2026-04-21", "EP-02", "Build embedder module",
     "Write rag/embedder.py wrapping Voyage AI. Handle batching, retries, rate limits. Add cost tracking.",
     "Task", "P0", 2, "To Do",
     "embed_text(str) and embed_batch(List[str]) work. Cost logged per call."),
    ("AG-010", 1, 2, "2026-04-21", "EP-02", "Build chunker module",
     "Write rag/chunker.py with strategies per content type: structured (company), paragraph (blog), email (one-per-email).",
     "Task", "P0", 1, "To Do",
     "Unit tests pass. Blog chunks respect 200-token size with 50-token overlap."),

    # Day 3 — Wed Apr 22 — Embedding Pipeline
    ("AG-011", 1, 3, "2026-04-22", "EP-02", "Ingest 50 sample company profiles",
     "Create target_companies.csv with 50 real ICP-matching companies. Run ingestion script to embed and store in company_profiles collection.",
     "Task", "P0", 2, "To Do",
     "50 companies visible in Qdrant. Metadata (industry, size, tech_stack) properly stored."),
    ("AG-012", 1, 3, "2026-04-22", "EP-02", "Ingest AnalyticsGear blog content",
     "Scrape all 6 blog posts from analyticsgear.com/insights. Chunk and embed into ag_content collection.",
     "Task", "P0", 1.5, "To Do",
     "Blog posts chunked correctly. Retrieval by topic returns relevant passages."),
    ("AG-013", 1, 3, "2026-04-22", "EP-02", "Ingest industry knowledge base",
     "Create 1-page summaries per target industry (Banking, Retail, Healthcare, SaaS, Manufacturing, Logistics). Embed into industry_knowledge collection.",
     "Task", "P1", 2, "To Do",
     "6 industry summaries indexed. Search by industry returns correct results."),

    # Day 4 — Thu Apr 23 — RAG Retrieval
    ("AG-014", 1, 4, "2026-04-23", "EP-02", "Build retriever module",
     "Write rag/retriever.py. Implement similarity search with metadata filtering. Support multi-collection queries.",
     "Task", "P0", 2, "To Do",
     "retrieve(query, collection, filters, top_k) returns ranked results with scores."),
    ("AG-015", 1, 4, "2026-04-23", "EP-02", "Build RAG context builder",
     "Write rag/context_builder.py. Given a prospect, orchestrate retrieval from multiple collections. Build injected-context prompt.",
     "Task", "P0", 2, "To Do",
     "build_context(prospect_id) returns combined context from company + ag_content + patterns."),
    ("AG-016", 1, 4, "2026-04-23", "EP-02", "Evaluate retrieval quality",
     "Write 20 test queries with expected results. Measure precision@5 and recall@5. Tune top-K and score thresholds.",
     "Task", "P1", 1.5, "To Do",
     "Precision@5 ≥ 80%. Results documented. Thresholds committed to config."),

    # Day 5 — Fri Apr 24 — Claude Integration
    ("AG-017", 1, 5, "2026-04-24", "EP-01", "Build Claude client wrapper",
     "Write llm/claude_client.py with tool use support, streaming, retries, cost tracking, and Langfuse tracing hooks.",
     "Task", "P0", 2, "To Do",
     "Client handles tool calls, streaming, errors gracefully. All calls traced."),
    ("AG-018", 1, 5, "2026-04-24", "EP-02", "Write first RAG-augmented email prompt",
     "Design prompt template: given prospect + retrieved context, generate personalized email. Test with 5 prospects.",
     "Task", "P0", 2, "To Do",
     "Generated emails reference company-specific details accurately. No hallucinated facts."),
    ("AG-019", 1, 5, "2026-04-24", "EP-02", "Validate email quality",
     "Manually review 10 AI-generated emails. Score: relevance, personalization, tone, CTA clarity. Tune prompt based on issues.",
     "QA", "P0", 1, "To Do",
     "8/10 emails rated 'would send as-is'. Common issues documented and fixed."),

    # Day 6 — Sat Apr 25 — First MCP Server (CRM)
    ("AG-020", 1, 6, "2026-04-25", "EP-03", "Set up Google Sheets CRM",
     "Create service account, credentials.json. Create spreadsheet with 3 sheets: Master Leads, Activity Log, Pipeline Dashboard. Apply schema from blueprint Section 7.2.",
     "Task", "P0", 1, "To Do",
     "Sheet accessible via service account. All columns present. Headers formatted."),
    ("AG-021", 1, 6, "2026-04-25", "EP-03", "Build mcp-crm server skeleton",
     "Create mcp_servers/mcp_crm/server.py. Implement server lifecycle (initialize, list_tools, call_tool).",
     "Task", "P0", 2, "To Do",
     "MCP server starts via stdio. Tool discovery returns empty list cleanly."),
    ("AG-022", 1, 6, "2026-04-25", "EP-03", "Implement CRM tools (read/write/log)",
     "Add tools: read_leads, write_lead, update_status, log_activity, get_pipeline_stats. Include proper JSON schemas.",
     "Task", "P0", 2.5, "To Do",
     "All 5 tools callable. Writes reflected in Google Sheet. Errors return meaningful messages."),
    ("AG-023", 1, 6, "2026-04-25", "EP-03", "Test mcp-crm with Claude Desktop",
     "Add server to Claude Desktop config. Manually test each tool. Verify discovery and execution.",
     "QA", "P0", 0.5, "To Do",
     "Claude Desktop successfully calls all 5 tools end-to-end."),

    # Day 7 — Sun Apr 26 — Sprint 1 Review
    ("AG-024", 1, 7, "2026-04-26", "EP-01", "Sprint 1 review & retrospective",
     "Document what worked, what didn't. Write Sprint 1 summary (learnings blog post draft #1 for later case study).",
     "Documentation", "P1", 1, "To Do",
     "Retro notes written. Blog post outline drafted."),
    ("AG-025", 1, 7, "2026-04-26", "EP-01", "Refactor & technical debt cleanup",
     "Address issues from week: rename confusing vars, add missing docstrings, improve error messages.",
     "Refactor", "P2", 2, "To Do",
     "Code passes linting. All public functions have docstrings."),
    ("AG-026", 1, 7, "2026-04-26", "EP-06", "Start email domain warmup",
     "Buy 'outreach.analyticsgear.com' domain. Set up SPF/DKIM/DMARC. Begin warmup via Instantly.ai (takes 3 weeks — start NOW).",
     "Task", "P0", 2, "To Do",
     "Domain purchased. DNS records verified. Warmup campaign running (5-10 emails/day)."),
    ("AG-027", 1, 7, "2026-04-26", "EP-01", "Plan Sprint 2 in detail",
     "Review Sprint 2 backlog. Adjust estimates based on Sprint 1 velocity. Identify dependencies.",
     "Planning", "P1", 1, "To Do",
     "Sprint 2 plan validated. Any scope changes committed."),

    # ========== SPRINT 2 (Week 2) ==========
    # Day 8 — Mon Apr 27 — MCP VectorDB + Email
    ("AG-028", 2, 8, "2026-04-27", "EP-03", "Build mcp-vectordb server",
     "Expose Qdrant as MCP tools: store_embedding, search_similar, get_by_id, update_metadata, delete_vector.",
     "Task", "P0", 2.5, "To Do",
     "All vector ops callable via MCP. Agents can retrieve/store through standard protocol."),
    ("AG-029", 2, 8, "2026-04-27", "EP-03", "Set up SendGrid account",
     "Create SendGrid account, verify sender domain (outreach.analyticsgear.com), configure webhooks for opens/clicks/replies.",
     "Task", "P0", 1, "To Do",
     "SendGrid verified. Webhook URL configured. Test email lands in inbox."),
    ("AG-030", 2, 8, "2026-04-27", "EP-03", "Build mcp-email server",
     "Implement tools: send_email, get_opens_clicks, check_deliverability, list_replies. Integrate with SendGrid API.",
     "Task", "P0", 3, "To Do",
     "Email send works end-to-end. Tracking events flow into CRM."),

    # Day 9 — Tue Apr 28 — MCP Scraper
    ("AG-031", 2, 9, "2026-04-28", "EP-03", "Build mcp-scraper base server",
     "Create mcp_servers/mcp_scraper/ with MCP skeleton. Add tools framework.",
     "Task", "P0", 1, "To Do",
     "Server starts. Tool registration works."),
    ("AG-032", 2, 9, "2026-04-28", "EP-03", "Add job board scraping tool",
     "Implement scrape_job_boards tool: LinkedIn Jobs + Indeed search for data engineering roles. Use Apify for reliability.",
     "Task", "P0", 2.5, "To Do",
     "Tool returns list of companies hiring data roles in last 7 days."),
    ("AG-033", 2, 9, "2026-04-28", "EP-03", "Add company enrichment tools",
     "Implement: get_tech_stack (BuiltWith/Wappalyzer), get_company_news (Google News), get_funding (Crunchbase).",
     "Task", "P0", 2, "To Do",
     "Given a domain, returns tech stack, recent news, funding status."),

    # Day 10 — Wed Apr 29 — MCP Apollo
    ("AG-034", 2, 10, "2026-04-29", "EP-03", "Set up Apollo.io account",
     "Create Apollo account (Basic tier $49). Configure ICP filters. Test API access.",
     "Task", "P0", 0.5, "To Do",
     "API key works. ICP search returns expected results."),
    ("AG-035", 2, 10, "2026-04-29", "EP-03", "Build mcp-apollo server",
     "Implement tools: search_companies, search_contacts, get_company_details, verify_email. Handle rate limits.",
     "Task", "P0", 3, "To Do",
     "All tools work. Rate limiting graceful. Results properly typed."),
    ("AG-036", 2, 10, "2026-04-29", "EP-03", "Test Apollo end-to-end",
     "Search 'CTOs in SaaS, 100-1000 employees, using Snowflake'. Verify 20 qualified results returned.",
     "QA", "P0", 1, "To Do",
     "Returns relevant leads. Emails verified. Metadata complete."),

    # Day 11 — Thu Apr 30 — MCP LinkedIn
    ("AG-037", 2, 11, "2026-04-30", "EP-03", "Set up LinkedIn automation",
     "Sign up for Phantombuster ($56/mo) or Expandi. Configure LinkedIn account connection (use warm account, 500+ connections).",
     "Task", "P0", 1, "To Do",
     "Phantombuster authenticated. Test connection request works manually."),
    ("AG-038", 2, 11, "2026-04-30", "EP-03", "Build mcp-linkedin server",
     "Implement: search_people, send_connection_request, send_message, get_profile, check_connection_status.",
     "Task", "P0", 3, "To Do",
     "All tools work via MCP. Rate limits respected (20 connections/day)."),
    ("AG-039", 2, 11, "2026-04-30", "EP-03", "Add anti-detection safeguards",
     "Implement human-like delays (random intervals), daily limits, business-hours-only operation.",
     "Task", "P1", 1, "To Do",
     "No rapid-fire actions. Account activity looks human."),

    # Day 12 — Fri May 1 — Base Agent
    ("AG-040", 2, 12, "2026-05-01", "EP-04", "Build base agent class",
     "Write agents/base_agent.py with ReAct loop: reason → act → observe → repeat. Handle tool calls from MCP.",
     "Task", "P0", 3, "To Do",
     "Agent executes goal with multi-step tool use. Logs each step."),
    ("AG-041", 2, 12, "2026-05-01", "EP-04", "Add agent observability",
     "Integrate Langfuse (or local JSON logging). Trace every reasoning step, tool call, and result.",
     "Task", "P0", 1.5, "To Do",
     "Every agent run produces structured trace. Can replay decision process."),
    ("AG-042", 2, 12, "2026-05-01", "EP-04", "Add error recovery & retries",
     "Handle: tool failures, LLM errors, rate limits, max-iterations. Graceful degradation.",
     "Task", "P0", 1.5, "To Do",
     "Agent survives injected failures. Stops at max-iterations with clear error."),

    # Day 13 — Sat May 2 — Prospecting Agent
    ("AG-043", 2, 13, "2026-05-02", "EP-04", "Write Prospecting Agent system prompt",
     "Craft system prompt defining role, ICP, workflow, tools, and output format. Iterate based on initial runs.",
     "Design", "P0", 2, "To Do",
     "Prompt committed to prompts/prospecting.txt. Agent follows workflow correctly."),
    ("AG-044", 2, 13, "2026-05-02", "EP-04", "Implement Prospecting Agent",
     "Build agents/prospecting_agent.py extending base agent. Wire up MCP servers: apollo, linkedin, scraper, vectordb, crm.",
     "Task", "P0", 2.5, "To Do",
     "Agent class instantiates. Discovers all expected tools."),
    ("AG-045", 2, 13, "2026-05-02", "EP-04", "Implement AI lead scoring",
     "Give agent ability to score 1-100 on 5 dimensions. Store scores in CRM and vector DB metadata.",
     "Task", "P0", 1.5, "To Do",
     "Scores correlate with manual assessment (> 75% agreement on sample of 20)."),

    # Day 14 — Sun May 3 — Sprint 2 Review
    ("AG-046", 2, 14, "2026-05-03", "EP-04", "End-to-end Prospecting Agent test",
     "Run agent with goal 'find 25 qualified leads'. Verify outputs, CRM writes, vector DB populations.",
     "QA", "P0", 2, "To Do",
     "25 leads in CRM. All scored. HOT leads flagged. No duplicates."),
    ("AG-047", 2, 14, "2026-05-03", "EP-04", "Tune prospecting workflow",
     "Based on test results, refine prompt, tool usage, scoring criteria. Re-run until quality hits target.",
     "QA", "P1", 1.5, "To Do",
     "Second test run produces higher-quality leads than first."),
    ("AG-048", 2, 14, "2026-05-03", "EP-01", "Sprint 2 retro & docs",
     "Write Sprint 2 retro. Draft blog post #2 on 'Building MCP Servers'.",
     "Documentation", "P1", 1, "To Do",
     "Retro notes done. Blog outline ready."),

    # ========== SPRINT 3 (Week 3) ==========
    # Day 15 — Mon May 4 — Outreach Agent Foundation
    ("AG-049", 3, 15, "2026-05-04", "EP-04", "Write Outreach Agent system prompt",
     "Design prompt for RAG-powered email generation. Define tone, length, structure, CTA rules.",
     "Design", "P0", 1.5, "To Do",
     "Prompt yields under-120-word emails with specific company references."),
    ("AG-050", 3, 15, "2026-05-04", "EP-04", "Implement Outreach Agent",
     "Build agents/outreach_agent.py. Connect to vectordb (for RAG), email, linkedin, crm MCP servers.",
     "Task", "P0", 2.5, "To Do",
     "Agent runs end-to-end. Drafts email using retrieved context."),
    ("AG-051", 3, 15, "2026-05-04", "EP-04", "Build email sequence templates",
     "Define HOT (5-email, 14-day) and WARM (3-email, 10-day) sequences in config/sequences.yaml.",
     "Design", "P0", 1.5, "To Do",
     "Sequences loaded by agent. Agent picks correct step based on sequence_step."),

    # Day 16 — Tue May 5 — Email Channel
    ("AG-052", 3, 16, "2026-05-05", "EP-06", "Verify email deliverability",
     "Test warmup progress via Mail-Tester. Check SPF/DKIM/DMARC. Verify not on blacklists (MXToolbox).",
     "QA", "P0", 1, "To Do",
     "Mail-Tester score ≥ 9/10. No blacklist hits."),
    ("AG-053", 3, 16, "2026-05-05", "EP-04", "Integrate Outreach Agent with email",
     "Wire up agent → email generation → send via mcp-email. Add daily limits (max 30/day initially).",
     "Task", "P0", 2, "To Do",
     "Agent sends real email. Tracking IDs logged. Opens/clicks tracked."),
    ("AG-054", 3, 16, "2026-05-05", "EP-04", "Add email safety guardrails",
     "Validate: no typos in lead name, no placeholder tokens left, no duplicate sends, allowlist for testing.",
     "Task", "P0", 1.5, "To Do",
     "Safety checks prevent bad sends. Test email with intentional errors is blocked."),

    # Day 17 — Wed May 6 — LinkedIn Channel
    ("AG-055", 3, 17, "2026-05-06", "EP-04", "Build LinkedIn message generation",
     "RAG-augmented connection notes (under 280 chars) and intro DMs. Tone matched to LinkedIn.",
     "Task", "P0", 2, "To Do",
     "Generated messages pass 'sounds human' test. Character limits respected."),
    ("AG-056", 3, 17, "2026-05-06", "EP-04", "Integrate Outreach Agent with LinkedIn",
     "Wire to mcp-linkedin. Handle connection flow: view profile → connect → message post-acceptance.",
     "Task", "P0", 2, "To Do",
     "Agent executes full LinkedIn sequence. Statuses tracked in CRM."),
    ("AG-057", 3, 17, "2026-05-06", "EP-04", "Multi-channel orchestration",
     "Agent decides channel mix per lead tier: HOT = email + LinkedIn + call-brief. WARM = email + LinkedIn.",
     "Task", "P0", 1.5, "To Do",
     "Agent executes correct channels based on tier. No channel skipped."),

    # Day 18 — Thu May 7 — Follow-up Agent
    ("AG-058", 3, 18, "2026-05-07", "EP-04", "Build engagement signal tracker",
     "Poll SendGrid webhooks + LinkedIn activity + CRM. Build unified engagement event stream.",
     "Task", "P0", 2, "To Do",
     "All engagement events logged within 5 min of occurrence."),
    ("AG-059", 3, 18, "2026-05-07", "EP-04", "Build reply classifier",
     "LLM-based classifier with 6 classes: INTERESTED, OBJECTION, NOT_NOW, NOT_INTERESTED, OUT_OF_OFFICE, WRONG_PERSON.",
     "Task", "P0", 2, "To Do",
     "Classifier achieves > 85% accuracy on labeled test set of 20 replies."),
    ("AG-060", 3, 18, "2026-05-07", "EP-04", "Write Follow-up Agent system prompt",
     "Define workflow: monitor → classify → draft response → update CRM → schedule next action.",
     "Design", "P0", 1, "To Do",
     "Prompt complete. Handles all 6 reply classes distinctly."),

    # Day 19 — Fri May 8 — Reply Handling
    ("AG-061", 3, 19, "2026-05-08", "EP-04", "Implement Follow-up Agent",
     "Build agents/followup_agent.py. Wire engagement tracker + classifier + response drafter.",
     "Task", "P0", 2.5, "To Do",
     "Agent processes 10 simulated replies correctly. Drafts appropriate responses."),
    ("AG-062", 3, 19, "2026-05-08", "EP-04", "Add human-in-loop for responses",
     "Before sending any auto-reply, flag for human review (email/Slack). Founder approves or edits.",
     "Task", "P0", 1.5, "To Do",
     "Draft replies land in review queue. Nothing sends without explicit approval."),
    ("AG-063", 3, 19, "2026-05-08", "EP-04", "Test full response cycle",
     "Simulate: inbound reply → classification → draft → approval → send → CRM update.",
     "QA", "P0", 1, "To Do",
     "Full cycle completes in under 5 minutes for each reply type."),

    # Day 20 — Sat May 9 — Memory System
    ("AG-064", 3, 20, "2026-05-09", "EP-05", "Build lead memory",
     "Per-lead interaction history stored in vector DB. Embed every email, reply, LinkedIn message.",
     "Task", "P0", 2, "To Do",
     "get_lead_history(lead_id) returns chronological interactions with context."),
    ("AG-065", 3, 20, "2026-05-09", "EP-05", "Build pattern memory",
     "Store successful outreach (got reply / booked meeting) in successful_patterns collection with metadata.",
     "Task", "P0", 1.5, "To Do",
     "Query 'emails that got replies from SaaS CTOs' returns ranked examples."),
    ("AG-066", 3, 20, "2026-05-09", "EP-05", "Memory-augment Outreach Agent",
     "Before drafting any follow-up, agent retrieves lead memory + successful patterns + uses as few-shot examples.",
     "Task", "P0", 2, "To Do",
     "Follow-up emails reference prior interactions accurately. No repetition."),

    # Day 21 — Sun May 10 — Sprint 3 Review
    ("AG-067", 3, 21, "2026-05-10", "EP-04", "End-to-end test of 3 agents",
     "Run Prospecting → Outreach → Follow-up in sequence. Verify handoffs and data flow.",
     "QA", "P0", 3, "To Do",
     "Full pipeline runs with 10 leads. All statuses correct at end. No data loss."),
    ("AG-068", 3, 21, "2026-05-10", "EP-05", "Bug fixes from integration testing",
     "Address any issues found. Retest.",
     "Bug", "P0", 1.5, "To Do",
     "All critical bugs resolved. Re-test passes."),
    ("AG-069", 3, 21, "2026-05-10", "EP-01", "Sprint 3 retro & docs",
     "Write retro. Draft blog post #3 on 'Multi-Agent Systems in Production'.",
     "Documentation", "P1", 1, "To Do",
     "Retro done. Blog outline complete."),

    # ========== SPRINT 4 (Week 4) ==========
    # Day 22 — Mon May 11 — Analytics Agent
    ("AG-070", 4, 22, "2026-05-11", "EP-04", "Build metrics calculator",
     "Compute: leads processed, emails sent, open rate, reply rate, meetings booked, cost per meeting, pipeline value.",
     "Task", "P0", 2, "To Do",
     "Metrics match CRM reality. Updated daily."),
    ("AG-071", 4, 22, "2026-05-11", "EP-04", "Write Analytics Agent system prompt",
     "Agent reviews weekly data, identifies trends, generates recommendations with specific A/B test suggestions.",
     "Design", "P0", 1, "To Do",
     "Prompt complete. Produces actionable recommendations."),
    ("AG-072", 4, 22, "2026-05-11", "EP-04", "Implement Analytics Agent",
     "Build agents/analytics_agent.py. Scheduled Friday afternoon. Sends report via email to founder.",
     "Task", "P0", 2, "To Do",
     "Weekly report generated. Sent automatically. Founder reads and acts."),

    # Day 23 — Tue May 12 — Feedback Loop
    ("AG-073", 4, 23, "2026-05-12", "EP-05", "Build engagement outcome tracker",
     "Track: which emails got replies, which lead profiles converted to meetings, which patterns worked.",
     "Task", "P0", 2, "To Do",
     "Outcome data flowing into feedback system daily."),
    ("AG-074", 4, 23, "2026-05-12", "EP-05", "Build pattern analyzer",
     "Weekly job: find commonalities in successful outreach. Auto-tag patterns and store.",
     "Task", "P0", 2, "To Do",
     "After simulated data, correctly identifies top-performing subject lines and CTAs."),
    ("AG-075", 4, 23, "2026-05-12", "EP-05", "Implement scoring optimizer",
     "Adjust lead scoring weights based on which lead profiles actually booked meetings. Log changes.",
     "Task", "P1", 1.5, "To Do",
     "Scoring weights change over time based on real data. Changes logged."),

    # Day 24 — Wed May 13 — Orchestrator
    ("AG-076", 4, 24, "2026-05-13", "EP-04", "Build PipelineOrchestrator",
     "Single entry point coordinating daily runs. Schedule: Prospect 6am, Outreach 8am, Follow-up 9am, Summary 6pm.",
     "Task", "P0", 2.5, "To Do",
     "Orchestrator runs full daily cycle. All 4 agents execute in correct order."),
    ("AG-077", 4, 24, "2026-05-13", "EP-04", "Add cross-agent error handling",
     "If Prospecting Agent fails, Outreach should still run on existing pipeline. No cascading failures.",
     "Task", "P0", 1.5, "To Do",
     "Injected failure in one agent doesn't break pipeline. Error notification sent."),
    ("AG-078", 4, 24, "2026-05-13", "EP-04", "Set up APScheduler / Prefect",
     "Wire orchestrator to scheduler. Test cron-like execution. Handle missed runs gracefully.",
     "Task", "P0", 1.5, "To Do",
     "Pipeline runs automatically at configured times. Missed runs caught up on next cycle."),

    # Day 25 — Thu May 14 — Production Setup
    ("AG-079", 4, 25, "2026-05-14", "EP-06", "Provision VPS",
     "Set up Hetzner CX22 or DigitalOcean ($10-20/mo). Configure SSH, firewall, fail2ban, basic hardening.",
     "Task", "P0", 1.5, "To Do",
     "VPS accessible. SSH-only auth. Non-root deploy user configured."),
    ("AG-080", 4, 25, "2026-05-14", "EP-06", "Dockerize all services",
     "Write Dockerfile for main app. docker-compose with Qdrant + app + Langfuse. Production secrets via env.",
     "Task", "P0", 2, "To Do",
     "docker compose up on VPS runs full stack. Services healthy."),
    ("AG-081", 4, 25, "2026-05-14", "EP-06", "Set up Langfuse monitoring",
     "Self-host Langfuse or use cloud. Configure traces for all agent runs and LLM calls.",
     "Task", "P0", 1.5, "To Do",
     "Every agent run visible in Langfuse. Costs, latency, errors tracked."),
    ("AG-082", 4, 25, "2026-05-14", "EP-06", "Set up alerting",
     "Email/Slack alerts on: agent failure, LLM errors > threshold, cost spike, no leads processed today.",
     "Task", "P0", 1, "To Do",
     "Test alert fires for each failure mode. Reaches founder within 2 min."),

    # Day 26 — Fri May 15 — Email Warmup & Final Checks
    ("AG-083", 4, 26, "2026-05-15", "EP-06", "Verify email warmup complete",
     "4 weeks in — should be fully warmed. Verify inbox placement via Mail-Tester and Gmail inbox check.",
     "QA", "P0", 0.5, "To Do",
     "Score ≥ 9/10. Emails land in Primary inbox, not Promotions/Spam."),
    ("AG-084", 4, 26, "2026-05-15", "EP-06", "Full pre-launch smoke test",
     "Run complete daily pipeline on staging data. Verify every component: agents, MCP servers, vector DB, email, tracking.",
     "QA", "P0", 3, "To Do",
     "All green. No errors. Metrics match expected."),
    ("AG-085", 4, 26, "2026-05-15", "EP-06", "Write operations runbook",
     "Document: how to check pipeline status, diagnose common failures, restart components, view logs, rollback.",
     "Documentation", "P0", 1.5, "To Do",
     "Runbook covers top 10 operational scenarios. Tested with self."),

    # Day 27 — Sat May 16 — Soft Launch
    ("AG-086", 4, 27, "2026-05-16", "EP-06", "Production deployment",
     "Deploy to VPS. Verify all services healthy. Pipeline scheduled but not yet sending.",
     "Task", "P0", 2, "To Do",
     "All services running on VPS. Schedules active. Dry-run mode on."),
    ("AG-087", 4, 27, "2026-05-16", "EP-06", "Soft launch with 10 real leads",
     "Disable dry-run. Let agents send real emails + LinkedIn messages to 10 carefully-reviewed leads.",
     "Task", "P0", 3, "To Do",
     "10 real outreach touches sent. All delivered. Tracking events received."),
    ("AG-088", 4, 27, "2026-05-16", "EP-06", "Monitor & fix production issues",
     "Watch logs, metrics, inbox. Fix any production-only issues that surface.",
     "Bug", "P0", 2, "To Do",
     "All issues triaged. P0/P1 resolved before scaling up."),

    # Day 28 — Sun May 17 — GO LIVE
    ("AG-089", 4, 28, "2026-05-17", "EP-06", "Scale to full daily volume",
     "Increase to full ICP targeting: 25 new leads/day, 30-50 emails/day, 20 LinkedIn connections/day.",
     "Task", "P0", 2, "To Do",
     "Pipeline running at target volume. All safety limits holding."),
    ("AG-090", 4, 28, "2026-05-17", "EP-06", "Build operations dashboard",
     "Simple dashboard (Streamlit or Langfuse) showing pipeline health, daily stats, cost, error rate.",
     "Task", "P1", 2, "To Do",
     "Dashboard accessible. Updates daily. Founder checks it each morning."),
    ("AG-091", 4, 28, "2026-05-17", "EP-06", "GO LIVE announcement + case study kickoff",
     "Post LinkedIn announcement. Start collecting case study metrics. Schedule weekly reviews.",
     "Milestone", "P0", 1, "To Do",
     "Public announcement posted. Metrics baseline captured. Weekly review scheduled."),
    ("AG-092", 4, 28, "2026-05-17", "EP-01", "Project retrospective",
     "Full 4-week retro. What went well, what didn't, key learnings. Draft case study outline.",
     "Documentation", "P1", 1.5, "To Do",
     "Retro complete. Case study outline ready to convert into blog posts and sales collateral."),
]

# Write tasks to sheet
for i, t in enumerate(TASKS, start=5):
    for j, v in enumerate(t, 1):
        cell = ws.cell(row=i, column=j, value=v)
        style_cell(cell, wrap=True, align="left" if j not in [1,2,3,8,9,10,11] else "center", size=9)

        if j == 1:  # Task ID
            style_cell(cell, bold=True, color=PURPLE, align="center", size=9)
        elif j == 2:  # Sprint
            style_cell(cell, bold=True, color=DARK_INDIGO, align="center", bg=LIGHT_INDIGO, size=9)
        elif j == 3:  # Day
            style_cell(cell, bold=True, color=GRAY, align="center", size=9)
        elif j == 5:  # Epic
            style_cell(cell, bold=True, color=TEAL, align="center", size=9)
        elif j == 6:  # Title
            style_cell(cell, bold=True, color="1E1E2E", align="left", size=9)
        elif j == 8:  # Type
            type_colors = {
                "Task": (INDIGO, LIGHT_INDIGO),
                "Design": (PURPLE, LIGHT_PURPLE),
                "QA": (TEAL, LIGHT_TEAL),
                "Bug": (RED, LIGHT_RED),
                "Refactor": (GRAY, LIGHT_GRAY),
                "Documentation": (ORANGE, LIGHT_ORANGE),
                "Planning": (ORANGE, LIGHT_ORANGE),
                "Milestone": (GREEN, LIGHT_GREEN),
            }
            if v in type_colors:
                fg, bg = type_colors[v]
                style_cell(cell, bold=True, color=fg, align="center", bg=bg, size=8)
        elif j == 9:  # Priority
            if v == "P0":
                style_cell(cell, bold=True, color=RED, align="center", bg=LIGHT_RED, size=9)
            elif v == "P1":
                style_cell(cell, bold=True, color=ORANGE, align="center", bg=LIGHT_ORANGE, size=9)
            else:
                style_cell(cell, bold=True, color=GRAY, align="center", size=9)
        elif j == 11:  # Status
            status_colors = {
                "To Do": (GRAY, LIGHT_GRAY),
                "In Progress": (ORANGE, LIGHT_ORANGE),
                "Done": (GREEN, LIGHT_GREEN),
                "Blocked": (RED, LIGHT_RED),
            }
            if v in status_colors:
                fg, bg = status_colors[v]
                style_cell(cell, bold=True, color=fg, align="center", bg=bg, size=9)

    # Row height based on description length
    ws.row_dimensions[i].height = 55

set_col_widths(ws, [10, 9, 6, 12, 9, 30, 55, 12, 9, 10, 12, 45])
ws.freeze_panes = "G5"  # Freeze top rows and first 6 columns

# ═══════════════════════════════════════════════════════════════════
#  SHEET 3: DAILY SCHEDULE
# ═══════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Daily Schedule")
add_title_row(ws, "Daily Schedule", "Day-by-day breakdown with hours, tasks, and focus areas", cols=7)

headers = ["Day", "Date", "Weekday", "Sprint", "Focus Area", "Tasks", "Est. Hours"]
for i, h in enumerate(headers, 1):
    style_header(ws.cell(row=4, column=i, value=h), bg=DARK_INDIGO)

# Aggregate tasks by day
from collections import defaultdict
days = defaultdict(lambda: {"sprint": 0, "date": "", "weekday": "", "tasks": [], "hours": 0, "focus": ""})

day_focus = {
    1: "Environment Setup",
    2: "Vector DB Foundation",
    3: "Embedding Pipeline",
    4: "RAG Retrieval",
    5: "Claude Integration",
    6: "First MCP Server (CRM)",
    7: "Sprint 1 Review + Start Email Warmup",
    8: "MCP VectorDB + Email",
    9: "MCP Scraper",
    10: "MCP Apollo",
    11: "MCP LinkedIn",
    12: "Base Agent Architecture",
    13: "Prospecting Agent",
    14: "Sprint 2 Review",
    15: "Outreach Agent Foundation",
    16: "Email Channel Integration",
    17: "LinkedIn Channel Integration",
    18: "Follow-up Agent",
    19: "Reply Handling",
    20: "Memory System",
    21: "Sprint 3 Review",
    22: "Analytics Agent",
    23: "Feedback Loop",
    24: "Orchestrator",
    25: "Production Setup",
    26: "Final Checks",
    27: "Soft Launch",
    28: "GO LIVE 🚀",
}

weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
start = date(2026, 4, 20)

for t in TASKS:
    day = t[2]
    days[day]["sprint"] = t[1]
    days[day]["date"] = t[3]
    d = start + timedelta(days=day-1)
    days[day]["weekday"] = weekdays[d.weekday()]
    days[day]["tasks"].append(f"{t[0]}: {t[5]}")
    days[day]["hours"] += t[9]
    days[day]["focus"] = day_focus.get(day, "")

for i, day_num in enumerate(sorted(days.keys()), start=5):
    d = days[day_num]
    tasks_text = "\n".join([f"• {t}" for t in d["tasks"]])

    cells_data = [
        (f"Day {day_num}", True, PURPLE, "center"),
        (d["date"], True, DARK_INDIGO, "center"),
        (d["weekday"], True, GRAY, "center"),
        (f"Sprint {d['sprint']}", True, DARK_INDIGO, "center"),
        (d["focus"], True, TEAL, "left"),
        (tasks_text, False, "1E1E2E", "left"),
        (f"{d['hours']:.1f} hrs", True, ORANGE if d["hours"] > 6 else GREEN, "center"),
    ]

    for j, (val, bold, color, align) in enumerate(cells_data, 1):
        c = ws.cell(row=i, column=j, value=val)
        style_cell(c, bold=bold, color=color, align=align, size=9)
        if d["weekday"] in ("Sat", "Sun"):
            c.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)

    # Row height based on number of tasks
    ws.row_dimensions[i].height = max(45, len(d["tasks"]) * 18)

set_col_widths(ws, [8, 12, 9, 10, 25, 80, 12])
ws.freeze_panes = "A5"

# ═══════════════════════════════════════════════════════════════════
#  SHEET 4: KANBAN BOARD
# ═══════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Kanban Board")
add_title_row(ws, "Kanban Board", "Backlog → To Do → In Progress → Review → Done", cols=5)

kanban_headers = ["📋 Backlog", "🟡 To Do (Current Sprint)", "🟠 In Progress", "🔵 Review", "✅ Done"]
kanban_colors = [GRAY, ORANGE, PURPLE, INDIGO, GREEN]
kanban_bgs = [LIGHT_GRAY, LIGHT_ORANGE, LIGHT_PURPLE, LIGHT_INDIGO, LIGHT_GREEN]

for i, (h, bg) in enumerate(zip(kanban_headers, kanban_colors), 1):
    c = ws.cell(row=4, column=i, value=h)
    style_header(c, bg=bg, size=12)
    ws.row_dimensions[4].height = 28

# Put Sprint 1 tasks in "To Do" column, rest in Backlog
for row_idx in range(5, 35):
    for col in range(1, 6):
        cell = ws.cell(row=row_idx, column=col, value="")
        cell.fill = PatternFill("solid", fgColor=kanban_bgs[col-1])
        cell.border = thin_border

# Add Sprint 1 tasks to "To Do" column
s1_tasks = [t for t in TASKS if t[1] == 1]
for i, t in enumerate(s1_tasks[:25], start=5):
    card_text = f"{t[0]}\n{t[5]}\nPriority: {t[8]}  •  {t[9]}h"
    c = ws.cell(row=i, column=2, value=card_text)
    style_cell(c, color=DARK_INDIGO, size=8, wrap=True, align="left", bg=WHITE)
    ws.row_dimensions[i].height = 55

# Backlog: Sprints 2-4
other_tasks = [t for t in TASKS if t[1] != 1]
for i, t in enumerate(other_tasks[:25], start=5):
    card_text = f"{t[0]} [S{t[1]}]\n{t[5][:60]}...\n{t[8]} • {t[9]}h"
    c = ws.cell(row=i, column=1, value=card_text)
    style_cell(c, color=GRAY, size=8, wrap=True, align="left", bg=WHITE)

set_col_widths(ws, [28, 30, 28, 28, 28])

# Note at bottom
ws.cell(row=36, column=1, value="💡 Move cards between columns as work progresses. The Sprint Backlog sheet is the source of truth for all tasks.")
ws.merge_cells(start_row=36, start_column=1, end_row=36, end_column=5)
ws.cell(row=36, column=1).font = Font(italic=True, color=GRAY, size=10)

# ═══════════════════════════════════════════════════════════════════
#  SHEET 5: MILESTONES
# ═══════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Milestones")
add_title_row(ws, "Project Milestones", "Key checkpoints with deliverables and success criteria", cols=5)

mile_headers = ["Milestone", "Target Date", "Day", "Deliverable", "Success Criteria"]
for i, h in enumerate(mile_headers, 1):
    style_header(ws.cell(row=4, column=i, value=h), bg=TEAL)

milestones = [
    ("M1 — Foundation Complete", "2026-04-24", "Day 5",
     "Working RAG pipeline with embedding, retrieval, and Claude integration",
     "Can generate a personalized email using retrieved company + blog context. Quality rated 8/10."),
    ("M2 — First MCP Server Live", "2026-04-25", "Day 6",
     "mcp-crm server operational and tested with Claude Desktop",
     "Claude Desktop can read/write CRM via MCP. Tool discovery works."),
    ("M3 — All MCP Servers Operational", "2026-05-01", "Day 12",
     "6 MCP servers: crm, vectordb, email, scraper, apollo, linkedin",
     "All servers run, all tools discoverable, all pass integration tests."),
    ("M4 — Prospecting Agent Live", "2026-05-03", "Day 14",
     "Autonomous agent finding 25 qualified leads/day",
     "Agent runs end-to-end without human intervention. 25 leads scored and stored."),
    ("M5 — Full Multi-Agent System", "2026-05-10", "Day 21",
     "Prospecting + Outreach + Follow-up agents integrated",
     "Full pipeline runs with 10 test leads. All statuses correctly updated."),
    ("M6 — Production Deployed", "2026-05-14", "Day 25",
     "System running on VPS with monitoring and scheduling",
     "Services healthy. Scheduled runs execute. Alerts fire on failure."),
    ("M7 — Soft Launch", "2026-05-16", "Day 27",
     "First 10 real outreach messages sent",
     "Real emails delivered. Tracking events received. Any P0 bugs fixed."),
    ("M8 — GO LIVE 🚀", "2026-05-17", "Day 28",
     "Full production volume — daily pipeline running autonomously",
     "25 leads/day, 30-50 emails/day, 20 LinkedIn connections/day. Dashboard live."),
]

for i, m in enumerate(milestones, start=5):
    for j, v in enumerate(m, 1):
        c = ws.cell(row=i, column=j, value=v)
        style_cell(c, wrap=True, align="left", size=10)
        if j == 1:
            style_cell(c, bold=True, color=TEAL, align="left", bg=LIGHT_TEAL, size=10)
        elif j in (2, 3):
            style_cell(c, bold=True, color=DARK_INDIGO, align="center", size=10)
    ws.row_dimensions[i].height = 52

set_col_widths(ws, [28, 15, 10, 45, 50])

# ═══════════════════════════════════════════════════════════════════
#  SHEET 6: DEPENDENCIES & CRITICAL PATH
# ═══════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Dependencies")
add_title_row(ws, "Dependencies & Critical Path", "Task dependencies and blocking relationships", cols=4)

dep_headers = ["Task", "Depends On", "Blocks", "Critical Path?"]
for i, h in enumerate(dep_headers, 1):
    style_header(ws.cell(row=4, column=i, value=h), bg=ORANGE)

dependencies = [
    ("AG-007 (Vector collections)", "AG-005 (Qdrant running)", "AG-008, AG-011, AG-012, AG-013", "YES"),
    ("AG-009 (Embedder)", "AG-003 (Voyage installed)", "AG-011, AG-012, AG-013, AG-014", "YES"),
    ("AG-014 (Retriever)", "AG-009, AG-011, AG-012", "AG-015, AG-018", "YES"),
    ("AG-018 (RAG email prompt)", "AG-014, AG-017 (Claude client)", "AG-049, AG-050 (Outreach Agent)", "YES"),
    ("AG-026 (Email warmup)", "None — START DAY 7", "AG-052, AG-083 (deliverability, go-live)", "YES (3-week critical path)"),
    ("AG-028-AG-038 (All MCP servers)", "AG-006 (Claude API)", "AG-040, AG-044 (Agents)", "YES"),
    ("AG-040 (Base agent)", "AG-017, MCP servers running", "AG-044, AG-050, AG-061, AG-072", "YES"),
    ("AG-044 (Prospecting Agent)", "AG-040, mcp-apollo, mcp-linkedin, mcp-vectordb, mcp-crm", "AG-046 (Sprint 2 goal)", "YES"),
    ("AG-050 (Outreach Agent)", "AG-044, mcp-email, AG-018 (RAG)", "AG-061 (Follow-up), AG-067 (E2E test)", "YES"),
    ("AG-061 (Follow-up Agent)", "AG-059 (Classifier), AG-058 (Signal tracker)", "AG-067 (E2E test)", "YES"),
    ("AG-064-AG-066 (Memory)", "AG-050 (Outreach sending emails)", "AG-073 (feedback), quality improvement", "Medium"),
    ("AG-076 (Orchestrator)", "All 4 agents operational", "AG-086 (Prod deploy)", "YES"),
    ("AG-079 (VPS)", "None — can run in parallel", "AG-080, AG-086 (deploy)", "YES"),
    ("AG-086 (Production deploy)", "AG-076, AG-080, AG-081, AG-082", "AG-087, AG-089 (launch)", "YES"),
    ("AG-083 (Warmup complete)", "AG-026 (started Day 7 = 4 weeks prior)", "AG-087 (real email sends)", "YES"),
    ("AG-089 (Scale to full volume)", "AG-087 (soft launch successful)", "🚀 GO LIVE", "YES"),
]

for i, d in enumerate(dependencies, start=5):
    for j, v in enumerate(d, 1):
        c = ws.cell(row=i, column=j, value=v)
        style_cell(c, wrap=True, align="left", size=9)
        if j == 4:
            color = RED if "YES" in v else GRAY
            bg = LIGHT_RED if "YES" in v else LIGHT_GRAY
            style_cell(c, bold=True, color=color, align="center", bg=bg, size=9)
    ws.row_dimensions[i].height = 35

set_col_widths(ws, [35, 40, 35, 20])

# ═══════════════════════════════════════════════════════════════════
#  SHEET 7: RISK REGISTER
# ═══════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Risk Register")
add_title_row(ws, "Risk Register", "Known risks with mitigation plans", cols=6)

risk_headers = ["Risk ID", "Risk", "Likelihood", "Impact", "Mitigation", "Owner"]
for i, h in enumerate(risk_headers, 1):
    style_header(ws.cell(row=4, column=i, value=h), bg=RED)

risks = [
    ("R-01", "Email domain gets blacklisted during warmup",
     "Medium", "HIGH",
     "Use separate domain from analyticsgear.com. Warm up slowly (5/day → 50/day over 3 weeks). Monitor MXToolbox weekly.",
     "Founder"),
    ("R-02", "LinkedIn account gets restricted due to automation",
     "High", "HIGH",
     "Strict 20-connections/day limit. Human-like delays. Business-hours only. Have backup account ready.",
     "Founder"),
    ("R-03", "MCP tool calls fail at scale (API rate limits)",
     "Medium", "Medium",
     "Implement exponential backoff. Cache responses where possible. Set per-API daily caps. Monitor in Langfuse.",
     "Founder"),
    ("R-04", "LLM costs spike unexpectedly",
     "Medium", "Medium",
     "Set monthly budget alerts in Anthropic console. Use Haiku for cheap tasks (scoring). Cache RAG context where possible.",
     "Founder"),
    ("R-05", "RAG retrieval quality too low (irrelevant results)",
     "Medium", "HIGH",
     "Evaluate retrieval in Day 4 (AG-016). Tune chunking, embedding model, top-K. Add re-ranker if needed.",
     "Founder"),
    ("R-06", "Agents hallucinate facts about prospects",
     "High", "HIGH",
     "Agent system prompts explicitly say 'only use facts from retrieved context'. Human review for first 2 weeks of outreach.",
     "Founder"),
    ("R-07", "Timeline slips due to complexity",
     "Medium", "Medium",
     "Buffer built into weekends. P2 tasks droppable. Keep scope — no feature creep. Weekly retros to catch early.",
     "Founder"),
    ("R-08", "Single point of failure (founder sick / traveling)",
     "Low", "HIGH",
     "Document everything as you build. Runbook on Day 26. Monitoring alerts ensure failures surface even if unattended.",
     "Founder"),
    ("R-09", "Apollo/LinkedIn API changes break scrapers",
     "Medium", "Medium",
     "MCP abstraction isolates changes. Pay for legitimate APIs where possible. Have fallback: manual CSV import.",
     "Founder"),
    ("R-10", "Real prospects complain about automated outreach",
     "Low", "HIGH",
     "Easy unsubscribe in every email. Honor GDPR/CAN-SPAM. Human approval for first 100 emails. Legitimate personalization (not spam).",
     "Founder"),
]

for i, r in enumerate(risks, start=5):
    for j, v in enumerate(r, 1):
        c = ws.cell(row=i, column=j, value=v)
        style_cell(c, wrap=True, align="left", size=9)
        if j == 1:
            style_cell(c, bold=True, color=RED, align="center", size=9)
        elif j in (3, 4):
            if v == "HIGH":
                style_cell(c, bold=True, color=RED, align="center", bg=LIGHT_RED, size=9)
            elif v == "Medium":
                style_cell(c, bold=True, color=ORANGE, align="center", bg=LIGHT_ORANGE, size=9)
            elif v == "Low":
                style_cell(c, bold=True, color=GREEN, align="center", bg=LIGHT_GREEN, size=9)
            elif v == "High":
                style_cell(c, bold=True, color=RED, align="center", bg=LIGHT_RED, size=9)
    ws.row_dimensions[i].height = 52

set_col_widths(ws, [8, 35, 12, 10, 60, 12])

# ═══════════════════════════════════════════════════════════════════
#  SHEET 8: BURN-DOWN / TRACKING
# ═══════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Progress Tracker")
add_title_row(ws, "Progress Tracker", "Daily progress log — update as you go", cols=7)

track_headers = ["Day", "Date", "Planned Tasks", "Completed", "Blocked/Issues", "Tomorrow's Priority", "Hours Spent"]
for i, h in enumerate(track_headers, 1):
    style_header(ws.cell(row=4, column=i, value=h), bg=INDIGO)

for day_num in sorted(days.keys()):
    d = days[day_num]
    row = day_num + 4
    style_cell(ws.cell(row=row, column=1, value=f"Day {day_num}"), bold=True, color=PURPLE, align="center")
    style_cell(ws.cell(row=row, column=2, value=d["date"]), bold=True, color=DARK_INDIGO, align="center")
    style_cell(ws.cell(row=row, column=3, value=len(d["tasks"])), align="center")
    style_cell(ws.cell(row=row, column=4, value=""), bg=LIGHT_GREEN, align="center")
    style_cell(ws.cell(row=row, column=5, value=""), bg=LIGHT_RED, align="left")
    style_cell(ws.cell(row=row, column=6, value=""), bg=LIGHT_YELLOW, align="left")
    style_cell(ws.cell(row=row, column=7, value=""), bg=LIGHT_INDIGO, align="center")
    ws.row_dimensions[row].height = 28

set_col_widths(ws, [8, 12, 14, 14, 35, 35, 14])

# Instructions row
ws.cell(row=34, column=1, value="💡 Fill in 'Completed', 'Blocked/Issues', and 'Tomorrow's Priority' at end of each day. Use as a daily standup log.")
ws.merge_cells(start_row=34, start_column=1, end_row=34, end_column=7)
ws.cell(row=34, column=1).font = Font(italic=True, color=GRAY, size=10)

# ─── Save ───
output = r"c:\analyticsgear\sales_pipeline\AnalyticsGear_AI_Pipeline_Project_Plan.xlsx"
wb.save(output)
print(f"Excel saved: {output}")
print(f"Total tasks: {len(TASKS)}")
print(f"Total hours: {sum(t[9] for t in TASKS):.1f}")
